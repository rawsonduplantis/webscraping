import openpyxl as xl

wb = xl.load_workbook('example.xlsx')

sheet1 = wb['Sheet1']
cellA1 = sheet1['A1']

print(cellA1.value)
print(type(cellA1.value))

print(cellA1.row)
print(cellA1.column)
print(cellA1.coordinate)

print(sheet1.cell(1,2).value)

print(sheet1.max_row)
print(sheet1.max_column)

for fruit in range(1, sheet1.max_row + 1):
    print(sheet1.cell(fruit, 2).value)
print()

print(xl.utils.get_column_letter(1))
print(xl.utils.get_column_letter(25))
print()
print(xl.utils.column_index_from_string(' Y'))

for currentrow in sheet1['A1':'C1']:
    print(currentrow)
    for currentcell in currentrow:
        print(currentcell.corodinate, currentcell.value)
        print()

for currentrow in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, max_col=sheet1.max_column):
    print(currentrow)
    print(currentrow[0].value)
    print(currentrow[1].value)
    print(currentrow[2].value)