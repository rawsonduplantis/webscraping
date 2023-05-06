# Find a 'scrappable' cryptocurrencies website where you can scrape the top 5 cryptocurrencies. Scrape the information and create a well formatted  Excel spreadsheet report. Be creative with your report (as in colors, spacing, currency, text size, etc). The report should display the name of the currency, the symbol (if applicable), the current price and % change in the last 24 hrs and corresponding price (based on % change)
# Furthermore, for Bitcoin and Ethereum, the program should alert you via text if the value increases or decrease within $5 of its current value.

# Webscrape and Excel
from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
# Twilio
from twilio.rest import Client
import keys

# Webscrape set-up
url = 'https://www.webull.com/quote/crypto'
# Request in case 404 Forbidden error
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url, headers=headers)
webpage = urlopen(req).read()
soup = BeautifulSoup(webpage, 'html.parser')
print(soup.title.text + '\n')

# Webscraping
priciest_tickers = ['BTCUSD', 'YFIUSD', 'ETHUSD', 'MKRUSD', 'BCHUSD']
ticker_info = []

crypto_data = soup.find_all('div', attrs={'class': 'table-row'})
for crypto in crypto_data:
    ticker = []
    for component in crypto:
        if str(component.text).__contains__(','):
            ticker.append(str(component.text).replace(',', ''))
        else:
            ticker.append(str(component.text))
    for pricey_ticker in priciest_tickers:
        if ticker[1].__contains__(pricey_ticker):
            ticker[1] = pricey_ticker
            ticker.pop(0)
            ticker.pop(len(ticker) - 1)
            ticker_info.append(ticker)

ticker_info.sort(key=lambda x: float(x[1]), reverse=True)
for ticker in ticker_info:
    print(ticker)

# Excel file
wb = xl.Workbook()
ws = wb.active

ws.title = 'Crypto Report'

ws['A1'] = 'Symbol'
ws['B1'] = 'Price ($USD)'
ws['C1'] = '% Change'
ws['D1'] = 'Change'
ws['E1'] = 'Open'
ws['F1'] = 'Prev Close'
ws['G1'] = 'High'
ws['H1'] = 'Low'

for i, crypto in enumerate(ticker_info):
    ws['A' + str(i + 2)] = crypto[0]
    ws['B' + str(i + 2)] = crypto[1]
    ws['C' + str(i + 2)] = crypto[2]
    ws['D' + str(i + 2)] = crypto[3]
    ws['E' + str(i + 2)] = crypto[4]
    ws['F' + str(i + 2)] = crypto[5]
    ws['G' + str(i + 2)] = crypto[6]
    ws['H' + str(i + 2)] = crypto[7]

header_font = Font(size=16, bold=True)

for cell in ws[1:1]:
    cell.font = header_font

for cell in ws['B':'B']:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws['D':'D']:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws['E':'E']:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws['F':'F']:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws['G':'G']:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws['H':'H']:
    cell.number_format = u'"$ "#,##0.00'

wb.save("CryptoReport.xlsx")

# Twilio

client = Client(keys.account_sid, keys.auth_token)

twilio_num = '+19379322599'
my_cell = '+12818258833'

for crypto in ticker_info:
    price_change = crypto[3]
    if abs(float(price_change)) > 5:
        text_message = client.messages.create(to=my_cell, from_=twilio_num, body=f"Price of {crypto[0]} has changed by {crypto[3]}.")