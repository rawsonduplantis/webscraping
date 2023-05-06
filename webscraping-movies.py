from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2023/'

page = urlopen(webpage)
soup = BeautifulSoup(page, 'html.parser')

print()
print(soup.title.text)
print()

table_rows = soup.findAll("tr")

#creating an excel file
wb = xl.Workbook()
ws = wb.active

ws.title = 'Box Office Report'

ws['A1'] = 'No.'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Gross'
ws['E1'] = 'Total Gross'
ws['F1'] = '% of Total Gross'

# Webscraping
No_Rank = ''
Movie_Title = ''
Release_Date = ''
Gross = 0.0
Total_Gross = 0.0
Percent_Total_Gross = 0.0

for x in range(1,6):
    td = table_rows[x].findAll("td")
    No_Rank = int(td[0].text)
    Movie_Title = td[1].text
    Release_Date = td[8].text
    Gross = int(td[5].text.replace(",","").replace("$",""))
    Total_Gross = int(td[7].text.replace(",","").replace("$",""))
    Precent_Total_Gross = round((Gross / Total_Gross) * 100, 2)

    #print(No_Rank, '', Movie_Title)
    #print('Released on:', Release_Date)
    #print()

    #applying to excel and iterating each row
    ws['A' + str(x+1)] = No_Rank
    ws['B' + str(x+1)] = Movie_Title
    ws['C' + str(x+1)] = Release_Date
    ws['D' + str(x+1)] = Gross
    ws['E' + str(x+1)] = Total_Gross
    ws['F' + str(x+1)] = str(Percent_Total_Gross) + '%'

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 25

header_font = Font(size=16, bold=True)

for cell in ws[1:1]:
    cell.font = header_font

wb.save("BoxOfficeReport.xlsx")